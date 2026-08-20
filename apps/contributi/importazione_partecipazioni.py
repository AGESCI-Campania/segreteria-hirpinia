"""Caricamento massivo delle partecipazioni da xlsx/CSV (D-21): unico punto
autorizzato a scrivere Partecipazione a partire da un file caricato in blocco.
View e comando di gestione chiamano solo `costruisci_piano_partecipazioni` e
`applica_piano_partecipazioni`, mai i modelli direttamente."""

from __future__ import annotations

import csv
import datetime
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from openpyxl import load_workbook

from apps.accounts.models import Utente
from apps.anagrafica.models import Capo
from apps.anagrafica.parser.buonacaccia import AVVISO, ERRORE, AnomaliaRiga
from apps.organizzazione.models import Gruppo

from .inserimento import risolvi_gruppo_competente
from .models import (
    Campagna,
    ImportazionePartecipazioni,
    Partecipazione,
    StatoCampagna,
    TipologiaCampo,
)

# Categorie di riga specifiche di D-21, distinte da ERRORE/AVVISO (già usati da
# M2/M3): AnomaliaRiga.livello è testo libero, estenderlo non tocca il modulo
# d'origine.
SOSPETTA = "sospetta"
CONFLITTO = "conflitto"

COLONNE_TRACCIATO = (
    "codice_socio",
    "cognome",
    "nome",
    "codice_tipologia",
    "data_inizio",
    "data_fine",
    "luogo",
    "quota_versata",
)


def leggi_righe_xlsx(contenuto: bytes) -> list[dict[str, Any]]:
    cartella = load_workbook(io.BytesIO(contenuto), read_only=True, data_only=True)
    foglio = cartella.active
    righe_iter = foglio.iter_rows(values_only=True)
    intestazione = [str(c).strip() if c is not None else "" for c in next(righe_iter, [])]
    return [
        dict(zip(intestazione, riga, strict=False))
        for riga in righe_iter
        if any(v not in (None, "") for v in riga)
    ]


def leggi_righe_csv(testo: str) -> list[dict[str, Any]]:
    return list(csv.DictReader(io.StringIO(testo)))


def _normalizza_data(valore: Any) -> datetime.date | None:
    if valore in (None, ""):
        return None
    if isinstance(valore, datetime.datetime):
        return valore.date()
    if isinstance(valore, datetime.date):
        return valore
    testo = str(valore).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(testo, formato).date()
        except ValueError:
            continue
    return None


def _normalizza_decimal(valore: Any) -> Decimal | None:
    if valore in (None, ""):
        return None
    if isinstance(valore, Decimal):
        return valore
    if isinstance(valore, int | float):
        return Decimal(str(valore))
    testo = str(valore).strip().replace(",", ".")
    if not testo:
        return None
    try:
        return Decimal(testo)
    except InvalidOperation:
        return None


@dataclass
class OperazionePartecipazione:
    numero_riga: int
    codice_socio: str
    gruppo: Gruppo
    tipologia: TipologiaCampo
    data_inizio: datetime.date
    data_fine: datetime.date
    luogo: str
    quota_versata: Decimal


@dataclass
class ConflittoPartecipazione:
    numero_riga: int
    codice_socio: str
    esistente: Partecipazione
    differenze: dict[str, tuple[Any, Any]]


@dataclass
class PianoPartecipazioni:
    campagna: Campagna
    valide: list[OperazionePartecipazione] = field(default_factory=list)
    gia_presenti: list[tuple[int, str]] = field(default_factory=list)
    conflitti: list[ConflittoPartecipazione] = field(default_factory=list)
    anomalie: list[AnomaliaRiga] = field(default_factory=list)

    @property
    def valido(self) -> bool:
        """Gate a livello di file (non di riga), D-21: il caricamento è
        consentito solo con campagna APERTA e dentro la finestra di
        inserimento."""
        return (
            self.campagna.stato == StatoCampagna.APERTA and self.campagna.in_finestra_inserimento()
        )


def costruisci_piano_partecipazioni(
    righe: list[dict[str, Any]], *, campagna: Campagna, utente: Utente
) -> PianoPartecipazioni:
    """Sola lettura: nessuna scrittura sul database."""
    piano = PianoPartecipazioni(campagna=campagna)
    if not piano.valido:
        piano.anomalie.append(
            AnomaliaRiga(
                0,
                ERRORE,
                "Campagna",
                f"La campagna {campagna.anno} non è aperta all'inserimento o è fuori "
                "dalla finestra prevista (D-21).",
                "",
            )
        )

    tipologie = {t.codice: t for t in TipologiaCampo.objects.filter(attiva=True)}

    for numero_riga, riga in enumerate(righe, start=2):
        codice_socio = str(riga.get("codice_socio") or "").strip()
        cognome_file = str(riga.get("cognome") or "").strip()
        nome_file = str(riga.get("nome") or "").strip()
        codice_tipologia = str(riga.get("codice_tipologia") or "").strip()
        luogo = str(riga.get("luogo") or "").strip()
        data_inizio = _normalizza_data(riga.get("data_inizio"))
        data_fine = _normalizza_data(riga.get("data_fine"))
        quota_versata = _normalizza_decimal(riga.get("quota_versata"))

        mancanti = [
            nome_campo
            for nome_campo, valore in (
                ("codice_socio", codice_socio),
                ("cognome", cognome_file),
                ("nome", nome_file),
                ("codice_tipologia", codice_tipologia),
                ("luogo", luogo),
            )
            if not valore
        ]
        if data_inizio is None:
            mancanti.append("data_inizio")
        if data_fine is None:
            mancanti.append("data_fine")
        if mancanti:
            piano.anomalie.append(
                AnomaliaRiga(
                    numero_riga,
                    ERRORE,
                    "Campi obbligatori",
                    f"Campi mancanti o non validi: {', '.join(mancanti)}.",
                    codice_socio,
                )
            )
            continue
        assert data_inizio is not None  # narrowing per mypy: già escluso da `mancanti` sopra
        assert data_fine is not None

        capo = Capo.objects.filter(codice_socio=codice_socio).first()
        if capo is None:
            piano.anomalie.append(
                AnomaliaRiga(
                    numero_riga, ERRORE, "Capo", "Codice socio assente in anagrafica.", codice_socio
                )
            )
            continue

        try:
            gruppo = risolvi_gruppo_competente(utente, codice_socio, anno_scout=campagna.anno)
        except (ValidationError, PermissionDenied) as exc:
            piano.anomalie.append(
                AnomaliaRiga(numero_riga, ERRORE, "Perimetro", str(exc), codice_socio)
            )
            continue

        if (
            cognome_file.casefold() != capo.cognome.strip().casefold()
            or nome_file.casefold() != capo.nome.strip().casefold()
        ):
            piano.anomalie.append(
                AnomaliaRiga(
                    numero_riga,
                    SOSPETTA,
                    "Controllo nome/cognome",
                    f"'{nome_file} {cognome_file}' non corrisponde all'anagrafica "
                    f"('{capo.nome} {capo.cognome}'): riga non importata.",
                    codice_socio,
                )
            )
            continue

        tipologia = tipologie.get(codice_tipologia)
        if tipologia is None:
            piano.anomalie.append(
                AnomaliaRiga(
                    numero_riga,
                    ERRORE,
                    "Tipologia",
                    f"Codice tipologia {codice_tipologia!r} inesistente o non attiva.",
                    codice_socio,
                )
            )
            continue

        quota = quota_versata if quota_versata is not None else tipologia.quota_default
        if quota is None:
            piano.anomalie.append(
                AnomaliaRiga(
                    numero_riga,
                    ERRORE,
                    "Quota",
                    "quota_versata assente e la tipologia non ha una quota di default.",
                    codice_socio,
                )
            )
            continue

        esistente = Partecipazione.objects.filter(
            campagna=campagna, capo_id=codice_socio, tipologia=tipologia, data_inizio=data_inizio
        ).first()
        if esistente is None:
            piano.valide.append(
                OperazionePartecipazione(
                    numero_riga=numero_riga,
                    codice_socio=codice_socio,
                    gruppo=gruppo,
                    tipologia=tipologia,
                    data_inizio=data_inizio,
                    data_fine=data_fine,
                    luogo=luogo,
                    quota_versata=quota,
                )
            )
            continue

        differenze: dict[str, tuple[Any, Any]] = {}
        if esistente.data_fine != data_fine:
            differenze["data_fine"] = (esistente.data_fine, data_fine)
        if esistente.luogo != luogo:
            differenze["luogo"] = (esistente.luogo, luogo)
        if esistente.quota_versata != quota:
            differenze["quota_versata"] = (esistente.quota_versata, quota)

        if not differenze:
            piano.gia_presenti.append((numero_riga, codice_socio))
        else:
            piano.conflitti.append(
                ConflittoPartecipazione(
                    numero_riga=numero_riga,
                    codice_socio=codice_socio,
                    esistente=esistente,
                    differenze=differenze,
                )
            )

    return piano


def applica_piano_partecipazioni(
    piano: PianoPartecipazioni, *, file_originale, utente: Utente | None
) -> ImportazionePartecipazioni:
    """Scrive solo `piano.valide` dentro un'unica transazione. Non tocca mai
    già-presenti/conflitti/sospette: un conflitto resta alla decisione
    dell'utente (D-21), mai auto-risolto."""
    if not piano.valido:
        raise ValueError(
            "Piano non applicabile: campagna non aperta o fuori finestra di inserimento."
        )

    with transaction.atomic():
        anomalie = list(piano.anomalie)
        conteggi = {
            "righe_valide": len(piano.valide),
            "partecipazioni_create": 0,
            "gia_presenti": len(piano.gia_presenti),
            "conflitti": len(piano.conflitti),
        }

        for operazione in piano.valide:
            partecipazione = Partecipazione(
                campagna=piano.campagna,
                capo_id=operazione.codice_socio,
                gruppo=operazione.gruppo,
                tipologia=operazione.tipologia,
                data_inizio=operazione.data_inizio,
                data_fine=operazione.data_fine,
                luogo=operazione.luogo,
                quota_versata=operazione.quota_versata,
                inserita_da=utente,
            )
            try:
                with transaction.atomic():
                    # exclude=["stato"]: vedi commento in campagne.py::apri_campagna.
                    partecipazione.full_clean(exclude=["stato"])
                    partecipazione.save()
            except ValidationError as exc:
                anomalie.append(
                    AnomaliaRiga(
                        operazione.numero_riga,
                        ERRORE,
                        "Partecipazione",
                        str(exc),
                        operazione.codice_socio,
                    )
                )
                continue
            conteggi["partecipazioni_create"] += 1

        anomalie_dict = [_anomalia_a_dict(a) for a in anomalie]
        anomalie_dict += [_conflitto_a_dict(c) for c in piano.conflitti]
        anomalie_dict += [
            {
                "numero_riga": n,
                "livello": AVVISO,
                "campo": "Duplicato",
                "dettaglio": "Già presente, non ricreata.",
                "codice_socio": cs,
            }
            for n, cs in piano.gia_presenti
        ]

        importazione = ImportazionePartecipazioni(
            campagna=piano.campagna, file=file_originale, utente=utente
        )
        importazione.save()
        importazione.conteggi = conteggi
        importazione.anomalie = anomalie_dict
        importazione.save(update_fields=["conteggi", "anomalie"])

    return importazione


def _anomalia_a_dict(a: AnomaliaRiga) -> dict:
    return {
        "numero_riga": a.numero_riga,
        "livello": a.livello,
        "campo": a.campo,
        "dettaglio": a.dettaglio,
        "codice_socio": a.codice_socio,
    }


def _conflitto_a_dict(c: ConflittoPartecipazione) -> dict:
    dettaglio = "; ".join(
        f"{campo}: {vecchio!r} -> {nuovo!r}" for campo, (vecchio, nuovo) in c.differenze.items()
    )
    return {
        "numero_riga": c.numero_riga,
        "livello": CONFLITTO,
        "campo": "Conflitto",
        "dettaglio": dettaglio,
        "codice_socio": c.codice_socio,
    }

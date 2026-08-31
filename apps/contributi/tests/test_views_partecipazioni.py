"""Test delle viste di contributi: perimetro e flusso anteprima→conferma di
D-21 (stesso principio di apps/anagrafica/tests/test_views_importazione.py)."""

import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from allauth.mfa.models import Authenticator
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.accounts.models import Delega, Ruolo, StatoUtente, TipoUtente, Utente
from apps.anagrafica.models import Capo, CensimentoCapo
from apps.contributi.models import Campagna, ImportazionePartecipazioni, TipologiaCampo
from apps.organizzazione.models import Gruppo

pytestmark = pytest.mark.django_db

ANNO = 2026


def _persona(email: str, **kwargs) -> Utente:
    n = Utente.objects.count()
    kwargs.setdefault("stato", StatoUtente.ATTIVO)
    return Utente.objects.create(username=f"u{n}", email=email, tipo=TipoUtente.PERSONA, **kwargs)


def _con_mfa_configurata(utente: Utente) -> Utente:
    Authenticator.objects.create(user=utente, type=Authenticator.Type.TOTP, data={"secret": "x"})
    return utente


@pytest.fixture
def gruppo() -> Gruppo:
    return Gruppo.objects.create(codice="E0133", nome="AVELLINO 1")


@pytest.fixture
def capo(gruppo) -> Capo:
    c = Capo.objects.create(codice_socio="10001", nome="MARIO", cognome="ROSSI")
    CensimentoCapo.objects.create(capo=c, anno_scout=ANNO, gruppo=gruppo)
    return c


@pytest.fixture
def campagna() -> Campagna:
    return Campagna.objects.create(
        anno=ANNO,
        budget=Decimal("1000.00"),
        data_inizio_inserimento=datetime.date(2025, 10, 1),
        data_fine_inserimento=datetime.date(2026, 9, 30),
    )


@pytest.fixture
def tipologia() -> TipologiaCampo:
    return TipologiaCampo.objects.get(codice="CFM")


@pytest.fixture
def cg_gruppo(gruppo) -> Utente:
    utente = _persona("cg@campania.agesci.it")
    Ruolo.objects.create(utente=utente, tipo=Ruolo.Tipo.CG, gruppo=gruppo)
    return _con_mfa_configurata(utente)


@pytest.fixture
def segreteria() -> Utente:
    utente = _persona("segreteria@campania.agesci.it")
    Ruolo.objects.create(utente=utente, tipo=Ruolo.Tipo.SEGRETERIA)
    return _con_mfa_configurata(utente)


def _file_xlsx(*, codice_socio="10001", cognome="ROSSI", nome="MARIO") -> SimpleUploadedFile:
    cartella = openpyxl.Workbook()
    foglio = cartella.active
    foglio.append(
        [
            "codice_socio",
            "cognome",
            "nome",
            "codice_tipologia",
            "data_inizio",
            "data_fine",
            "luogo",
            "quota_versata",
        ]
    )
    foglio.append(
        [codice_socio, cognome, nome, "CFM", "01/06/2026", "08/06/2026", "Base scout", ""]
    )
    buffer = BytesIO()
    cartella.save(buffer)
    return SimpleUploadedFile(
        "partecipazioni.xlsx",
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class TestPerimetro:
    def test_accesso_negato_senza_ruolo(self, client, campagna):
        utente = _persona("senza-ruolo@campania.agesci.it")
        client.force_login(utente)
        response = client.get(f"/contributi/campagne/{campagna.pk}/")
        assert response.status_code == 403

    def test_accesso_negato_anonimo(self, client, campagna):
        response = client.get(f"/contributi/campagne/{campagna.pk}/")
        assert response.status_code == 302

    def test_cg_accede_al_dettaglio_campagna(self, client, campagna, cg_gruppo):
        client.force_login(cg_gruppo)
        response = client.get(f"/contributi/campagne/{campagna.pk}/")
        assert response.status_code == 200


class TestCampagnaCrea:
    def test_segreteria_diretta_crea_campagna(self, client, segreteria):
        client.force_login(segreteria)
        response = client.post(
            "/contributi/campagne/nuova/",
            {
                "anno": 2027,
                "budget": "1000.00",
                "tetto_per_partecipazione": "50.00",
                "data_inizio_inserimento": "2026-10-01",
                "data_fine_inserimento": "2027-09-30",
            },
        )
        assert response.status_code == 302
        assert Campagna.objects.filter(anno=2027).exists()

    def test_delegato_di_segreteria_non_crea_campagna(self, client, segreteria):
        # D-11: i delegati sono esclusi dai parametri di campagna.
        delegato = _persona("delegato@campania.agesci.it")
        _con_mfa_configurata(delegato)
        ruolo = Ruolo.objects.get(utente=segreteria)
        Delega.objects.create(
            delegante=segreteria,
            delegato=delegato,
            ruolo=ruolo,
            data_fine=datetime.date.today() + datetime.timedelta(days=30),
        )
        client.force_login(delegato)
        response = client.post(
            "/contributi/campagne/nuova/",
            {
                "anno": 2027,
                "budget": "1000.00",
                "tetto_per_partecipazione": "50.00",
                "data_inizio_inserimento": "2026-10-01",
                "data_fine_inserimento": "2027-09-30",
            },
        )
        assert response.status_code == 403
        assert not Campagna.objects.filter(anno=2027).exists()

    def test_cg_non_crea_campagna(self, client, cg_gruppo):
        client.force_login(cg_gruppo)
        response = client.get("/contributi/campagne/nuova/")
        assert response.status_code == 403


class TestFlussoAnteprimaConferma:
    def test_anteprima_non_scrive_nulla(self, client, campagna, capo, tipologia, cg_gruppo):
        client.force_login(cg_gruppo)
        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/importa/",
            {"file": _file_xlsx()},
        )

        assert response.status_code == 200
        assert ImportazionePartecipazioni.objects.count() == 0

    def test_conferma_dopo_anteprima_scrive(self, client, campagna, capo, tipologia, cg_gruppo):
        client.force_login(cg_gruppo)
        client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/importa/",
            {"file": _file_xlsx()},
        )

        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/importa/conferma/"
        )

        assert response.status_code == 302
        assert ImportazionePartecipazioni.objects.count() == 1

    def test_conferma_senza_anteprima_non_scrive(self, client, campagna, cg_gruppo):
        client.force_login(cg_gruppo)
        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/importa/conferma/"
        )

        assert response.status_code == 302
        assert ImportazionePartecipazioni.objects.count() == 0

    def test_conferma_con_sessione_di_altro_utente_non_scrive(
        self, client, campagna, capo, tipologia, cg_gruppo, gruppo
    ):
        client.force_login(cg_gruppo)
        client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/importa/",
            {"file": _file_xlsx()},
        )

        altro_cg = _persona("altro-cg@campania.agesci.it")
        Ruolo.objects.create(utente=altro_cg, tipo=Ruolo.Tipo.CG, gruppo=gruppo)
        _con_mfa_configurata(altro_cg)
        client.force_login(altro_cg)

        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/importa/conferma/"
        )

        assert response.status_code == 302
        assert ImportazionePartecipazioni.objects.count() == 0

    def test_modello_xlsx_scaricabile(self, client, campagna, segreteria):
        client.force_login(segreteria)
        response = client.get(f"/contributi/campagne/{campagna.pk}/partecipazioni/modello.xlsx")

        assert response.status_code == 200
        cartella = openpyxl.load_workbook(BytesIO(response.content))
        intestazione = [c.value for c in next(cartella.active.iter_rows())]
        assert intestazione == [
            "codice_socio",
            "cognome",
            "nome",
            "codice_tipologia",
            "data_inizio",
            "data_fine",
            "luogo",
            "quota_versata",
        ]


class TestPartecipazioneInserisciErroriCampo:
    """Issue GitHub #1: gli errori di validazione devono comparire sul
    campo corretto (non genericamente in cima al form) e senza i codici di
    decisione interni (issue #4)."""

    def test_data_inizio_fuori_finestra_su_campo_corretto(
        self, client, campagna, capo, tipologia, cg_gruppo
    ):
        client.force_login(cg_gruppo)
        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/inserisci/",
            {
                "codice_socio": capo.codice_socio,
                "tipologia": tipologia.pk,
                "descrizione_altro": "",
                "data_inizio": "2020-01-01",
                "data_fine": "2020-01-08",
                "luogo": "Base scout",
                "quota_versata": "10.00",
                "note": "",
            },
        )

        assert response.status_code == 200
        form = response.context["form"]
        assert "data_inizio" in form.errors
        assert "__all__" not in form.errors
        messaggio = form.errors["data_inizio"][0]
        assert "01/10/2025" in messaggio
        assert "30/09/2026" in messaggio
        assert "D-10" not in messaggio

    def test_descrizione_altro_mancante_su_campo_corretto(self, client, campagna, capo, cg_gruppo):
        tipologia_altro = TipologiaCampo.objects.get(codice="ALTRO")
        client.force_login(cg_gruppo)
        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/inserisci/",
            {
                "codice_socio": capo.codice_socio,
                "tipologia": tipologia_altro.pk,
                "descrizione_altro": "",
                "data_inizio": "2026-06-01",
                "data_fine": "2026-06-08",
                "luogo": "Base scout",
                "quota_versata": "10.00",
                "note": "",
            },
        )

        assert response.status_code == 200
        form = response.context["form"]
        assert "descrizione_altro" in form.errors
        assert "__all__" not in form.errors

    def test_codice_socio_fuori_perimetro_resta_errore_generale(
        self, client, campagna, tipologia, cg_gruppo
    ):
        # ValidationError "piatta" da risolvi_gruppo_competente (nessun
        # dict per campo): deve restare un errore non-field (nessuna
        # regressione).
        client.force_login(cg_gruppo)
        response = client.post(
            f"/contributi/campagne/{campagna.pk}/partecipazioni/inserisci/",
            {
                "codice_socio": "99999",
                "tipologia": tipologia.pk,
                "descrizione_altro": "",
                "data_inizio": "2026-06-01",
                "data_fine": "2026-06-08",
                "luogo": "Base scout",
                "quota_versata": "10.00",
                "note": "",
            },
        )

        assert response.status_code == 200
        form = response.context["form"]
        assert "__all__" in form.errors
        assert "data_inizio" not in form.errors
